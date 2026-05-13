from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()
ws = wb.active
ws.title = "Affiliate Applications"

headers = [
    "Done",
    "Tier",
    "Tool",
    "Network",
    "Commission",
    "Cookie / duration",
    "Why / fit",
    "Wiki page?",
    "Applied date",
    "Status",
    "Notes",
]

rows = [
    # Tier 1 — already have a wiki page
    ["", 1, "Gamma", "PartnerStack", "30% first year", "12 months", "Presentation AI; existing /tools/gamma/ page", "Yes", "", "", ""],
    ["", 1, "Lindy", "PartnerStack", "30% Y1 + 15% Y2", "Recurring", "AI assistant; existing /tools/lindy/ page", "Yes", "", "", ""],
    ["", 1, "Consensus", "PartnerStack", "30% / ~$36 per signup", "12 months", "Academic search; existing /tools/consensus/ page", "Yes", "", "", ""],
    ["", 1, "Beautiful.ai", "PartnerStack", "30% Pro / $43 per referral", "Annual / first month", "Presentation AI; existing /tools/beautiful-ai/ page; pairs with Gamma comparison", "Yes", "", "", ""],
    ["", 1, "Activepieces", "PartnerStack", "Up to $17.5k/yr per enterprise", "3 years", "No-code automation; fills n8n / Make gap (both declined)", "Yes", "", "", ""],

    # Tier 2 — strong AI fit, write page after approval
    ["", 2, "Reclaim.ai", "PartnerStack", "40% recurring + $1/signup", "12 months", "AI calendar — no competitor on wiki yet; highest recurring %", "No (build)", "", "", ""],
    ["", 2, "Castmagic", "PartnerStack", "40% recurring", "Lifetime recurring", "AI audio→content for creators", "No (build)", "", "", ""],
    ["", 2, "MeetGeek", "PartnerStack", "30% recurring", "Lifetime", "Rival to Otter/Fireflies (both in your Tier 1 list elsewhere)", "No (build)", "", "", ""],
    ["", 2, "Tidio", "PartnerStack", "Up to 30% + $50 first sale", "Recurring", "Lyro AI agent + chat; fits ai-chatbots category", "No (build)", "", "", ""],
    ["", 2, "SaneBox", "PartnerStack", "Up to 30% recurring", "Recurring", "AI email productivity; converts well per program desc", "No (build)", "", "", ""],
    ["", 2, "Synthflow AI", "PartnerStack", "20%", "12 months", "Voice agents — complements ElevenLabs", "No (build)", "", "", ""],
    ["", 2, "Browse AI", "PartnerStack", "20% forever", "Lifetime", "No-code web scraping; recurring lifetime is rare", "No (build)", "", "", ""],
    ["", 2, "MindStudio", "PartnerStack", "20% revenue share", "12 months", "No-code AI agent builder", "No (build)", "", "", ""],
    ["", 2, "Pinecone", "PartnerStack", "10%", "12 months", "Vector DB — high-search-volume AI infra term", "No (build)", "", "", ""],
    ["", 2, "BLACKBOX AI", "PartnerStack", "30%", "12 months", "AI coding assistant; Copilot competitor", "No (build)", "", "", ""],
    ["", 2, "Pangram Labs", "PartnerStack", "30% recurring", "12 months", "AI-detection; constant search demand, no page yet", "No (build)", "", "", ""],

    # Tier 3 — decent fit, lower priority
    ["", 3, "Vista Social", "PartnerStack", "25-50% recurring + 90d cookie", "12 months", "Social mgmt; 90d cookie is unusually long", "No", "", "", ""],
    ["", 3, "Brand24", "PartnerStack", "20% recurring", "Recurring", "Media monitoring (AI-tagged)", "No", "", "", ""],
    ["", 3, "Reply.io", "PartnerStack", "15% lifetime on Jason AI SDR", "Lifetime on Jason", "Sales engagement; Jason AI SDR is the AI angle", "No", "", "", ""],
    ["", 3, "Databox", "PartnerStack", "20% / up to $1,200 per customer", "12 months", "Analytics with AI insights", "No", "", "", ""],
    ["", 3, "Lusha", "PartnerStack", "Up to 20%", "12 months", "B2B contact data (tangential to AI)", "No", "", "", ""],
    ["", 3, "Demodesk", "PartnerStack", "20% (~$1K per referral avg)", "12 months", "Sales AI / conversation intel", "No", "", "", ""],
    ["", 3, "Flowith", "PartnerStack", "Up to 40%", "TBD", "AI creativity / multi-thread agents; unknown traction", "No", "", "", ""],
    ["", 3, "Logome.ai", "PartnerStack", "50% per sale", "Per sale", "Logo gen; high % but low search volume", "No", "", "", ""],
    ["", 3, "Typewise", "PartnerStack", "20% recurring (up to $20k/customer)", "Recurring", "CS agent platform; enterprise ACV is the upside", "No", "", "", ""],
    ["", 3, "TheTop", "PartnerStack", "50% Y1", "12 months", "AI chief of staff; unproven product but high % if it converts", "No", "", "", ""],
]

# Header row
header_fill = PatternFill("solid", start_color="1F4E78")
header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
thin = Side(style="thin", color="C0C0C0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col_idx, h in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

# Data rows
tier_fills = {
    1: PatternFill("solid", start_color="D9EAD3"),
    2: PatternFill("solid", start_color="FFF2CC"),
    3: PatternFill("solid", start_color="F4F4F4"),
}

for r_offset, row in enumerate(rows, start=2):
    tier = row[1]
    for c_offset, val in enumerate(row, start=1):
        cell = ws.cell(row=r_offset, column=c_offset, value=val)
        cell.font = Font(name="Arial", size=11)
        cell.alignment = left_wrap if c_offset in (7, 11) else Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = border
        if c_offset == 2:
            cell.fill = tier_fills[tier]
            cell.alignment = center

# Column widths
widths = {
    "A": 7,   # Done
    "B": 6,   # Tier
    "C": 18,  # Tool
    "D": 13,  # Network
    "E": 32,  # Commission
    "F": 18,  # Cookie / duration
    "G": 52,  # Why / fit
    "H": 12,  # Wiki page?
    "I": 13,  # Applied date
    "J": 14,  # Status
    "K": 32,  # Notes
}
for letter, w in widths.items():
    ws.column_dimensions[letter].width = w

# Row heights
ws.row_dimensions[1].height = 28
for r in range(2, len(rows) + 2):
    ws.row_dimensions[r].height = 32

# Freeze panes (header + tool col)
ws.freeze_panes = "C2"

# Data validation
done_dv = DataValidation(type="list", formula1='"☐,✅"', allow_blank=True)
done_dv.add(f"A2:A{len(rows) + 1}")
ws.add_data_validation(done_dv)

status_dv = DataValidation(type="list", formula1='"Not started,Applied,Approved,Declined,On hold"', allow_blank=True)
status_dv.add(f"J2:J{len(rows) + 1}")
ws.add_data_validation(status_dv)

# Pre-fill Done cells with the empty checkbox character
for r in range(2, len(rows) + 2):
    ws.cell(row=r, column=1, value="☐").alignment = center
    ws.cell(row=r, column=10, value="Not started").alignment = Alignment(horizontal="left", vertical="center")

# Auto-filter
ws.auto_filter.ref = f"A1:K{len(rows) + 1}"

# Second sheet — quick legend / how to use
ws2 = wb.create_sheet("Legend")
ws2["A1"] = "How to use this sheet"
ws2["A1"].font = Font(bold=True, size=14, name="Arial")
notes = [
    "",
    "Import into Google Sheets via File → Import → Upload → Replace spreadsheet.",
    "",
    "Tiers:",
    "  Tier 1 (green) — apply now; wiki page already exists on aipedia.wiki.",
    "  Tier 2 (yellow) — apply, then build the wiki page after approval.",
    "  Tier 3 (grey) — decent fit, lower priority.",
    "",
    "Columns:",
    "  Done — click the cell, pick ☐ or ✅ from the dropdown.",
    "  Status — Not started / Applied / Approved / Declined / On hold.",
    "  Applied date — fill in once submitted (YYYY-MM-DD).",
    "",
    "After each application, also update src/data/_meta/tools-registry.json for the matching tool entry.",
]
for i, line in enumerate(notes, start=2):
    c = ws2.cell(row=i, column=1, value=line)
    c.font = Font(name="Arial", size=11)
ws2.column_dimensions["A"].width = 90

wb.save("affiliate-applications.xlsx")
print("Saved affiliate-applications.xlsx with", len(rows), "rows")
